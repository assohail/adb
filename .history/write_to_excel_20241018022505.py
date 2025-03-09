import sys
from openpyxl import Workbook
import os

# Get the tap count from the command line argument
tap_count = int(sys.argv[1]) if len(sys.argv) > 1 else 0

# Define the filename
filename = "tap_count.xlsx"

# Check if the file already exists
if os.path.exists(filename):
    # Load existing workbook
    workbook = Workbook(filename)
else:
    # Create a new workbook
    workbook = Workbook()

# Select the active worksheet
sheet = workbook.active

# Add headers if the sheet is empty
if sheet.max_row == 1:
    sheet.append(["Total Taps"])

# Append the tap count
sheet.append([tap_count])

# Save the workbook
workbook.save(filename)
