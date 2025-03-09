import sys
from openpyxl import Workbook, load_workbook
import os

# Get the tap count from the command line argument
tap_count = int(sys.argv[1]) if len(sys.argv) > 1 else 0

# Define the filename
filename = "tap_count.xlsx"

# Check if the file already exists
if os.path.exists(filename):
    # Load existing workbook
    workbook = load_workbook(filename)
    sheet = workbook.active if workbook.sheetnames else workbook.create_sheet("Sheet1")
else:
    # Create a new workbook and add a default sheet
    workbook = Workbook()
    sheet = workbook.active

# Add headers if the sheet is empty
if sheet.max_row == 1 and sheet['A1'].value is None:  # Check if the first cell is empty
    sheet.append(["Total Taps"])

# Append the tap count
sheet.append(['below with no filter'])

# Save the workbook
workbook.save(filename)
print(f"Tap data saved to {filename}.")
